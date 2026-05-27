#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate MOE SERPS POC Acceptance Test Excel
Two sheets: Chinese + English
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ─── Colors ───────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1E3A5F"
C_HEADER_FG  = "FFFFFF"
C_SECTION_BG = "2E75B6"
C_SECTION_FG = "FFFFFF"
C_ALT_ROW    = "EBF3FB"
C_BORDER     = "BFBFBF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def wrap(h="left"):
    return Alignment(wrap_text=True, horizontal=h, vertical="top")

def set_header(ws, row, col, value, bg=C_HEADER_BG):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    c.alignment = center()
    c.border = thin_border()

def set_body(ws, row, col, value, alt=False, bold=False, h="left"):
    c = ws.cell(row=row, column=col, value=value)
    if alt:
        c.fill = fill(C_ALT_ROW)
    c.font = Font(name="Calibri", bold=bold, size=10)
    c.alignment = wrap(h=h)
    c.border = thin_border()

def set_section(ws, row, ncols, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = fill(C_SECTION_BG)
    c.font = Font(name="Calibri", bold=True, size=11, color=C_SECTION_FG)
    c.alignment = wrap(h="left")
    c.border = thin_border()

# ─── Account data ─────────────────────────────────────────────────────────────
ACCOUNTS = [
    ("admin",      "System Admin",                "admin123",    "系统管理员 / Admin"),
    ("principal",  "Hjh Rashidah Binti Mohamad",  "principal123","校长 / Principal"),
    ("hod01",      "Dr. Azman Bin Ishak",          "hod123",      "部门主任 / HOD"),
    ("manager",    "Hj Kamaruddin",                "Demo@2026",   "学务管理 / Manager"),
    ("finance",    "Finance Officer",              "finance123",  "财务 / Finance"),
    ("admissions", "Admissions Officer",           "Demo@2026",   "招生 / Admissions"),
    ("drsiti",     "Dr. Siti Nurhaliza",           "Demo@2026",   "教师 / Teacher"),
    ("teacher01",  "Ms. Aminah Binti Hassan",      "teacher123",  "教师 / Teacher"),
    ("student001", "Ahmad Bin Abdullah",           "student123",  "学生 / Student"),
    ("adam",       "Adam Bin Haris",               "Demo@2026",   "学生 / Student"),
    ("parent01",   "Hj Abdullah Bin Mahmud",       "parent123",   "家长 / Parent"),
    ("fatimah",    "Fatimah Binti Yusof",          "Demo@2026",   "家长 / Parent"),
]

# ─── Test cases ───────────────────────────────────────────────────────────────
# (module_zh, module_en, feature_zh, feature_en, steps_zh, steps_en, expected_zh, expected_en, account)

TESTS = [
    # ── 登录 / Login ──────────────────────────────────────────────────────────
    (
        "登录与权限", "Login & Access",
        "管理员登录", "Admin Login",
        "1. 打开系统地址\n2. 用户名 admin，密码 admin123\n3. 点击 [登录]",
        "1. Open the system URL\n2. Enter username: admin, password: admin123\n3. Click Login",
        "进入管理员仪表板，侧边栏显示全部菜单（含系统设置）",
        "Admin dashboard loads; sidebar shows full menu including System Settings",
        "admin / admin123",
    ),
    (
        "登录与权限", "Login & Access",
        "校长登录 - 角色隔离", "Principal Login - Role Isolation",
        "1. 注销当前账号\n2. 用户名 principal，密码 principal123\n3. 点击 [登录]",
        "1. Log out\n2. Enter username: principal, password: principal123\n3. Click Login",
        "进入校长仪表板；左侧菜单不含 [系统设置]；可见 KPI 卡片与学校概览",
        "Principal dashboard loads; sidebar does NOT include System Settings; KPI cards and school overview visible",
        "principal / principal123",
    ),
    (
        "登录与权限", "Login & Access",
        "教师登录", "Teacher Login",
        "1. 用户名 drsiti，密码 Demo@2026\n2. 点击 [登录]",
        "1. Enter username: drsiti, password: Demo@2026\n2. Click Login",
        "进入教师仪表板，显示我的班级、今日课表、待提交报告等",
        "Teacher dashboard loads showing My Classes, Today's Timetable, Pending Reports",
        "drsiti / Demo@2026",
    ),
    (
        "登录与权限", "Login & Access",
        "学生登录", "Student Login",
        "1. 用户名 student001，密码 student123\n2. 点击 [登录]",
        "1. Enter username: student001, password: student123\n2. Click Login",
        "进入学生视图，显示个人成绩、课表、考勤；无管理功能",
        "Student view loads showing personal grades, timetable, attendance; no admin features",
        "student001 / student123",
    ),
    (
        "登录与权限", "Login & Access",
        "家长登录", "Parent Login",
        "1. 用户名 parent01，密码 parent123\n2. 点击 [登录]",
        "1. Enter username: parent01, password: parent123\n2. Click Login",
        "进入家长视图，显示子女信息（成绩、考勤、通知）",
        "Parent view loads showing child's grades, attendance, and notices",
        "parent01 / parent123",
    ),
    (
        "登录与权限", "Login & Access",
        "密码错误拦截", "Wrong Password Blocked",
        "1. 任意用户名，密码故意填错\n2. 点击 [登录]",
        "1. Enter any username with incorrect password\n2. Click Login",
        "显示错误提示，拒绝登录",
        "Error message shown; access denied",
        "任意 / Any",
    ),
    (
        "登录与权限", "Login & Access",
        "语言切换（英/中/马来）", "Language Switch (EN / ZH / MS)",
        "1. 登录后找到右上角语言切换\n2. 依次切换 English -> 中文 -> Bahasa Melayu",
        "1. After login, find language switcher in top-right\n2. Switch: English -> Zhong Wen -> Bahasa Melayu",
        "页面所有文字（菜单、按钮、标签）随之切换；刷新后保持选择",
        "All UI text (menus, buttons, labels) switches instantly; preference retained after page refresh",
        "admin / admin123",
    ),
    # ── 仪表板 / Dashboard ────────────────────────────────────────────────────
    (
        "仪表板", "Dashboard",
        "管理员总览 KPI", "Admin KPI Cards",
        "1. 以 admin 登录\n2. 打开仪表板首页",
        "1. Login as admin\n2. Open Dashboard",
        "4 张 KPI 卡片：在校学生数、教职工数、本月出勤率、本月收款；数据来自数据库",
        "4 KPI cards: Total Students, Total Staff, Monthly Attendance, Monthly Revenue; all from database",
        "admin / admin123",
    ),
    (
        "仪表板", "Dashboard",
        "校长视图 - 学校概览", "Principal Overview",
        "1. 以 principal 登录\n2. 打开仪表板",
        "1. Login as principal\n2. Open Dashboard",
        "看到学生总数、教职工数、班级数、出勤趋势折线图",
        "Displays student count, staff count, class count, and attendance trend chart",
        "principal / principal123",
    ),
    (
        "仪表板", "Dashboard",
        "教师仪表板", "Teacher Dashboard",
        "1. 以 teacher01 登录\n2. 打开仪表板",
        "1. Login as teacher01\n2. Open Dashboard",
        "显示我的班级、今日课表、待批改作业数、出勤异常提醒",
        "Shows My Classes, today's timetable, pending assignments, attendance alerts",
        "teacher01 / teacher123",
    ),
    (
        "仪表板", "Dashboard",
        "HOD 仪表板", "HOD Dashboard",
        "1. 以 hod01 登录\n2. 打开仪表板",
        "1. Login as hod01\n2. Open Dashboard",
        "显示部门教师绩效摘要、CPD 完成率、待审批事项",
        "Shows department teacher performance summary, CPD completion rate, pending approvals",
        "hod01 / hod123",
    ),
    (
        "仪表板", "Dashboard",
        "高风险学生预警", "At-Risk Student Alert",
        "1. 以 principal 登录\n2. 打开高风险学生模块或仪表板相应区块",
        "1. Login as principal\n2. Open At-Risk Students module or dashboard section",
        "展示风险评分 Top N 学生列表，含风险因素说明（出勤低、成绩差等）",
        "Lists top-N students by risk score with risk factor explanations (low attendance, poor grades)",
        "principal / principal123",
    ),
    # ── SIS ──────────────────────────────────────────────────────────────────
    (
        "SIS 学生信息", "SIS",
        "查看学生列表", "Student List",
        "1. 以 admin 登录\n2. 导航至 SIS -> 学生管理",
        "1. Login as admin\n2. Navigate to SIS -> Student Management",
        "分页展示学生列表，可按姓名/班级搜索；每行有查看详情按钮",
        "Paginated student list; searchable by name/class; each row has View Details button",
        "admin / admin123",
    ),
    (
        "SIS 学生信息", "SIS",
        "查看学生详情", "Student Profile",
        "1. 在学生列表中点击任意学生名称或 [详情] 按钮",
        "1. Click any student name or the Details button in the list",
        "打开学生档案页：基本信息、监护人、成绩、出勤、财务记录",
        "Student profile opens: basic info, guardian, grades, attendance, finance records",
        "admin / admin123",
    ),
    (
        "SIS 学生信息", "SIS",
        "新生入学向导（4 步）", "Enrollment Wizard (4 Steps)",
        "1. 点击 [新增学生]\n2. 依次填写：个人信息 -> 监护人 -> 入学信息 -> 确认\n3. 提交",
        "1. Click New Student\n2. Complete wizard: Personal Info -> Guardian -> Enrollment -> Confirm\n3. Submit",
        "系统根据出生日期自动推算年级；提交后学生出现在列表中；各步骤有输入验证",
        "System auto-calculates grade from DOB; student appears in list after submit; each step validates",
        "admissions / Demo@2026",
    ),
    (
        "SIS 学生信息", "SIS",
        "表单必填校验", "Required Field Validation",
        "1. 打开新生入学向导\n2. 不填写姓名，直接点 [下一步]",
        "1. Open Enrollment Wizard\n2. Leave Name blank and click Next",
        "姓名输入框下方显示红色错误提示，无法进入下一步",
        "Red error message shown below name field; cannot proceed without required fields",
        "admissions / Demo@2026",
    ),
    # ── EMS ──────────────────────────────────────────────────────────────────
    (
        "EMS 员工管理", "EMS",
        "查看教职工列表", "Staff List",
        "1. 以 admin 登录\n2. 导航至 EMS -> 员工管理",
        "1. Login as admin\n2. Navigate to EMS -> Staff Management",
        "展示所有教职工，含姓名、职位、部门、状态列",
        "Lists all staff with name, position, department, and status columns",
        "admin / admin123",
    ),
    (
        "EMS 员工管理", "EMS",
        "CPD 培训记录", "CPD Training Records",
        "1. 以 hod01 登录\n2. 进入 EMS -> CPD 培训记录\n3. 为某教师添加一条培训记录",
        "1. Login as hod01\n2. Go to EMS -> CPD Records\n3. Add a training record for a teacher",
        "记录保存成功；该教师 CPD 完成小时数更新；可按年度筛选",
        "Record saved; teacher's CPD hours updated; filterable by academic year",
        "hod01 / hod123",
    ),
    (
        "EMS 员工管理", "EMS",
        "绩效评估工作流", "Performance Evaluation Workflow",
        "1. 以 teacher01 登录，完成自评表单并提交\n2. 切换 hod01 登录，审批该评估",
        "1. Login as teacher01, complete self-assessment and submit\n2. Switch to hod01, approve the evaluation",
        "教师侧：提交后状态变为 [待审批]\nHOD 侧：可见待审批列表；审批后状态变为 [已完成]",
        "Teacher side: status changes to Pending Approval\nHOD side: approval list visible; after approval status = Completed",
        "teacher01 / teacher123\nhod01 / hod123",
    ),
    # ── SMS ──────────────────────────────────────────────────────────────────
    (
        "SMS 学校管理", "SMS",
        "查看自动生成课表", "Auto-Generated Timetable",
        "1. 以 teacher01 或 admin 登录\n2. 进入 SMS -> 课表管理",
        "1. Login as teacher01 or admin\n2. Go to SMS -> Timetable",
        "展示按班级或教师维度的周课表；时段、科目、教室可见",
        "Weekly timetable shown by class or teacher; time slots, subjects, rooms visible",
        "teacher01 / teacher123",
    ),
    (
        "SMS 学校管理", "SMS",
        "设施预约", "Facility Booking",
        "1. 以 drsiti 登录\n2. 进入 SMS -> 设施预约\n3. 选择 [礼堂]，选日期时段，提交",
        "1. Login as drsiti\n2. Go to SMS -> Facility Booking\n3. Select Hall, choose date/time, submit",
        "预约申请提交成功；在预约列表中可见该条记录；时段冲突时提示错误",
        "Booking submitted successfully; record appears in booking list; conflict triggers an error",
        "drsiti / Demo@2026",
    ),
    (
        "SMS 学校管理", "SMS",
        "学校日历", "School Calendar",
        "1. 以任意账号登录\n2. 进入学校日历页面",
        "1. Login with any account\n2. Navigate to School Calendar page",
        "Ant Design 日历显示当月事件（假期、考试、活动）；可切换月份",
        "Ant Design calendar shows current month events (holidays, exams, activities); month navigation works",
        "admin / admin123",
    ),
    # ── 财务 / Finance ────────────────────────────────────────────────────────
    (
        "财务", "Finance",
        "财务概览", "Finance Overview",
        "1. 以 finance 登录\n2. 打开财务模块",
        "1. Login as finance\n2. Open Finance module",
        "显示收款统计、支出汇总、逾期账单数等 KPI；图表可交互",
        "KPI cards: total revenue, expenses, overdue bills; interactive charts",
        "finance / finance123",
    ),
    (
        "财务", "Finance",
        "查看学费记录", "Student Fee Records",
        "1. 以 finance 登录\n2. 进入学费/收款管理\n3. 搜索某学生",
        "1. Login as finance\n2. Go to Fees / Payments\n3. Search for a student",
        "展示该学生的缴费历史及未缴金额",
        "Payment history and outstanding balance shown for the student",
        "finance / finance123",
    ),
    # ── 招生 / Admissions ─────────────────────────────────────────────────────
    (
        "招生", "Admissions",
        "招生申请列表", "Application List",
        "1. 以 admissions 登录\n2. 打开招生模块",
        "1. Login as admissions\n2. Open Admissions module",
        "列出所有入学申请，含状态（待审、已批、已拒）",
        "Lists all enrollment applications with status (Pending, Approved, Rejected)",
        "admissions / Demo@2026",
    ),
    # ── 通知 / Notifications ──────────────────────────────────────────────────
    (
        "通知系统", "Notifications",
        "铃铛通知徽章", "Bell Notification Badge",
        "1. 以任意账号登录\n2. 查看页面右上角铃铛图标",
        "1. Login with any account\n2. Check bell icon in top-right",
        "铃铛显示未读数徽章；点击展开通知列表；内容与账号角色相关",
        "Bell shows unread count badge; click to expand notification list; content relevant to user role",
        "admin / admin123",
    ),
    (
        "通知系统", "Notifications",
        "SMTP 邮件配置（管理员）", "SMTP Email Config (Admin)",
        "1. 以 admin 登录\n2. 进入 系统设置 -> AI/邮件配置\n3. 查看 SMTP 配置界面",
        "1. Login as admin\n2. Go to System Settings -> AI / Email Config\n3. Review SMTP section",
        "显示 SMTP Host、端口、用户名等输入项；有 [测试发送] 按钮；密码字段脱敏",
        "SMTP Host, Port, Username fields visible; Test Send button available; password field masked",
        "admin / admin123",
    ),
    # ── AI 助手 / AI Assistant ────────────────────────────────────────────────
    (
        "AI 助手", "AI Assistant",
        "AI 对话（流式响应）", "Streaming Chat Response",
        "1. 以任意账号登录\n2. 点击 AI 助手入口\n3. 输入问题（如：本校本月出勤率是多少），发送",
        "1. Login with any account\n2. Open AI Assistant\n3. Type a question (e.g. What is the attendance rate this month?) and send",
        "AI 以流式方式逐字返回答复；回答与学校数据相关；无网络错误",
        "AI responds with streaming text word by word; response is school-data relevant; no network error",
        "admin / admin123",
    ),
    # ── 移动端 / Mobile H5 ────────────────────────────────────────────────────
    (
        "移动端 H5", "Mobile H5",
        "学生移动门户", "Student Portal (Mobile)",
        "1. 手机或浏览器移动模式打开移动端地址\n2. 以 student001 / student123 登录",
        "1. Open mobile URL on phone or browser mobile mode\n2. Login as student001 / student123",
        "移动端界面友好；可查看成绩、课表、通知；底部导航栏正常",
        "Mobile-friendly UI loads; grades, timetable, notifications accessible; bottom nav works",
        "student001 / student123",
    ),
    (
        "移动端 H5", "Mobile H5",
        "家长移动门户", "Parent Portal (Mobile)",
        "1. 以 parent01 / parent123 在移动端登录",
        "1. Login as parent01 / parent123 on mobile",
        "可查看子女成绩、考勤、通知；界面适配手机屏幕",
        "Child's grades, attendance, and notices accessible; UI fits mobile screen",
        "parent01 / parent123",
    ),
    (
        "移动端 H5", "Mobile H5",
        "PWA 离线缓存", "PWA Offline Cache",
        "1. 手机浏览器打开移动端并登录\n2. 断开网络\n3. 刷新页面",
        "1. Open mobile URL on phone and login\n2. Disconnect network\n3. Refresh page",
        "已加载的页面可离线访问；显示离线提示而非空白页",
        "Previously loaded page stays accessible offline; offline indicator shown instead of blank page",
        "任意移动端 / Any mobile",
    ),
    # ── 系统 / System ─────────────────────────────────────────────────────────
    (
        "系统", "System",
        "404 错误页面", "404 Error Page",
        "1. 在地址栏输入不存在的路径，如 /nonexistent",
        "1. Type a non-existent path in the browser, e.g. /nonexistent",
        "显示友好 404 页面，包含 [返回首页] 按钮；文字经过 i18n",
        "Friendly 404 page shown with Back to Home button; text is i18n-translated",
        "无需登录 / No login required",
    ),
    (
        "系统", "System",
        "403 无权限页面", "403 Forbidden Page",
        "1. 以低权限账号登录（如 student001）\n2. 手动访问管理员专属路径",
        "1. Login as low-privilege account (e.g. student001)\n2. Manually access an admin-only path",
        "显示 403 无权限页面；含返回按钮；不暴露系统内部错误",
        "403 Forbidden page shown with Back button; no internal system errors exposed",
        "student001 / student123",
    ),
]

# ─── Sheet builder ────────────────────────────────────────────────────────────
def build_sheet(wb, sheet_title, lang):
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_view.showGridLines = False

    ZH = (lang == "zh")
    today = date.today().strftime("%Y-%m-%d")

    if ZH:
        main_title   = "MOE SERPS POC — 功能验收测试表"
        sub_title    = "文莱教育部学校企业资源规划系统  |  验收日期：" + today
        url_line     = "  PC 端地址：http://localhost:5173     |     移动端地址：http://localhost:5174"
        acct_section = "  测试账号一览"
        acct_hdrs    = ["用户名", "姓名", "密码", "角色"]
        test_section = "  功能验收测试用例"
        test_hdrs    = ["#", "模块", "功能点", "操作步骤", "预期结果", "测试账号", "结果\n(通过/失败)", "备注"]
        footer       = "说明：所有测试项以演示数据库数据为准。请在\"结果\"列填写 Pass 或 Fail，并在备注栏记录问题。"
        idx_module   = 0  # zh module index in TESTS tuple
        idx_feature  = 2
        idx_steps    = 4
        idx_expected = 6
    else:
        main_title   = "MOE SERPS POC — Functional Acceptance Test Checklist"
        sub_title    = "Ministry of Education Brunei — School ERP System  |  Test Date: " + today
        url_line     = "  System URL (PC): http://localhost:5173     |     Mobile URL: http://localhost:5174"
        acct_section = "  Demo Accounts"
        acct_hdrs    = ["Username", "Display Name", "Password", "Role"]
        test_section = "  Functional Test Cases"
        test_hdrs    = ["#", "Module", "Feature", "Steps", "Expected Result", "Test Account", "Result\n(Pass/Fail)", "Notes"]
        footer       = "Note: All test items are based on the production demo database. Record Pass or Fail in the Result column and note any issues in Notes."
        idx_module   = 1  # en module index
        idx_feature  = 3
        idx_steps    = 5
        idx_expected = 7

    # Column widths
    col_w = [5, 18, 22, 42, 40, 22, 13, 22]
    NCOLS = len(col_w)
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # ── Title ──────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 34
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value=main_title)
    c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    c.fill = fill("1E3A5F")
    c.alignment = center()
    r += 1

    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value=sub_title)
    c.font = Font(name="Calibri", italic=True, size=10, color="FFFFFF")
    c.fill = fill("2C4770")
    c.alignment = center()
    r += 1

    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value=url_line)
    c.font = Font(name="Calibri", size=10, color="1E3A5F")
    c.fill = fill("D9E8F5")
    c.alignment = Alignment(horizontal="left", vertical="center")
    r += 2

    # ── Account table ──────────────────────────────────────────────────────
    set_section(ws, r, NCOLS, acct_section)
    ws.row_dimensions[r].height = 22
    r += 1

    for ci, h in enumerate(acct_hdrs, 1):
        set_header(ws, r, ci, h, bg="1E5FA0")
    for ci in range(len(acct_hdrs)+1, NCOLS+1):
        ws.cell(row=r, column=ci).fill = fill("1E5FA0")
        ws.cell(row=r, column=ci).border = thin_border()
    ws.row_dimensions[r].height = 22
    r += 1

    for i, (uname, dname, pwd, role) in enumerate(ACCOUNTS):
        alt = (i % 2 == 0)
        set_body(ws, r, 1, uname, alt=alt)
        set_body(ws, r, 2, dname, alt=alt)
        set_body(ws, r, 3, pwd, alt=alt, h="center")
        set_body(ws, r, 4, role, alt=alt)
        for ci in range(5, NCOLS+1):
            cell = ws.cell(row=r, column=ci)
            if alt:
                cell.fill = fill(C_ALT_ROW)
            cell.border = thin_border()
        ws.row_dimensions[r].height = 16
        r += 1
    r += 1

    # ── Test cases table ───────────────────────────────────────────────────
    set_section(ws, r, NCOLS, test_section)
    ws.row_dimensions[r].height = 22
    r += 1

    for ci, h in enumerate(test_hdrs, 1):
        set_header(ws, r, ci, h, bg="1E5FA0")
    ws.row_dimensions[r].height = 32
    r += 1

    current_module = None
    case_num = 0
    for row_data in TESTS:
        module   = row_data[idx_module]
        feature  = row_data[idx_feature]
        steps    = row_data[idx_steps]
        expected = row_data[idx_expected]
        account  = row_data[8]

        if module != current_module:
            current_module = module
            set_section(ws, r, NCOLS, "    " + module)
            ws.row_dimensions[r].height = 20
            r += 1

        case_num += 1
        alt = (case_num % 2 == 0)

        set_body(ws, r, 1, str(case_num), alt=alt, h="center")
        set_body(ws, r, 2, module, alt=alt)
        set_body(ws, r, 3, feature, alt=alt, bold=True)
        set_body(ws, r, 4, steps, alt=alt)
        set_body(ws, r, 5, expected, alt=alt)
        set_body(ws, r, 6, account, alt=alt, h="center")

        # Result cell
        rc = ws.cell(row=r, column=7, value="")
        rc.fill = fill("F2FFF2")
        rc.border = thin_border()
        rc.alignment = center()

        # Notes cell
        nc = ws.cell(row=r, column=8, value="")
        nc.border = thin_border()
        nc.alignment = Alignment(wrap_text=True, vertical="top")

        step_lines = steps.count("\n") + 1
        exp_lines  = expected.count("\n") + 1
        ws.row_dimensions[r].height = max(step_lines, exp_lines) * 15 + 8
        r += 1

    r += 1

    # ── Footer ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value=footer)
    c.font = Font(name="Calibri", italic=True, size=9, color="555555")
    c.fill = fill("F5F5F5")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = thin_border()
    ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A4"

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]

    build_sheet(wb, "中文验收测试", "zh")
    build_sheet(wb, "EN Acceptance Test", "en")

    out_path = "doc/MOE_SERPS_Acceptance_Test.xlsx"
    wb.save(out_path)
    print("Saved:", out_path)
    print("Sheets:", wb.sheetnames)
    print("Test cases:", len(TESTS))

if __name__ == "__main__":
    main()
